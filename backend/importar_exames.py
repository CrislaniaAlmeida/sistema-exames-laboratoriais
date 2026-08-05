"""
Script de importacao em massa de exames e laboratorios de apoio.
Le as planilhas em dados_importacao/ e insere no banco de dados
apontado pela variavel DATABASE_URL (local ou producao).

Uso:
    python importar_exames.py local       -> importa no banco local (.env)
    python importar_exames.py producao    -> importa no banco de producao (Neon)
"""

import sys
import os
import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from dotenv import load_dotenv

DATABASE_URL_PRODUCAO = os.getenv("DATABASE_URL_PRODUCAO")

PASTA_DADOS = os.path.join(os.path.dirname(__file__), "dados_importacao")
ARQUIVO_APOIO = os.path.join(PASTA_DADOS, "APOIO - ATUALIZADA.xlsx")
ARQUIVO_MNEMONICOS = os.path.join(PASTA_DADOS, "TABELA_MNEMONICOS_HIAS_.xlsx")


def escolher_banco():
    load_dotenv()

    if len(sys.argv) < 2 or sys.argv[1] not in ("local", "producao"):
        print("Uso: python importar_exames.py [local|producao]")
        sys.exit(1)

    if sys.argv[1] == "producao":
        print("⚠️  Voce esta prestes a importar dados no banco de PRODUCAO (Neon).")
        confirmacao = input("Digite 'CONFIRMAR' para continuar: ")
        if confirmacao != "CONFIRMAR":
            print("Cancelado.")
            sys.exit(0)
        return create_engine(os.getenv("DATABASE_URL_PRODUCAO"))
    else:
        return create_engine(os.getenv("DATABASE_URL"))


def carregar_mnemonicos():
    """Retorna um dicionario {nome_exame_maiusculo: sigla}"""
    wb = openpyxl.load_workbook(ARQUIVO_MNEMONICOS, data_only=True)
    ws = wb["Lista Completa"]
    mnemonicos = {}
    for linha in ws.iter_rows(min_row=4, values_only=True):
        nome, sigla = linha[0], linha[1]
        if nome and sigla:
            mnemonicos[nome.strip().upper()] = sigla.strip()
    return mnemonicos


def carregar_apoio():
    """
    Retorna:
    - lista de nomes de laboratorios (unicos)
    - lista de dicts de exames (um por nome de exame unico, primeira ocorrencia)
    """
    wb = openpyxl.load_workbook(ARQUIVO_APOIO, data_only=True)
    ws = wb["Página1"]

    laboratorios = set()
    exames = {}

    for linha in ws.iter_rows(min_row=2, values_only=True):
        nome_lab, contrato, procedimento, qtd_contratada = linha[0], linha[1], linha[2], linha[3]
        saldo = linha[19] if len(linha) > 19 else None

        if not nome_lab or not procedimento:
            continue

        nome_lab = nome_lab.strip()
        nome_exame = procedimento.strip()
        laboratorios.add(nome_lab)

        if nome_exame.upper() not in exames:
            exames[nome_exame.upper()] = {
                "nome": nome_exame,
                "laboratorio_nome": nome_lab,
                "quantidade_contratada": int(qtd_contratada) if qtd_contratada else None,
                "quantidade_restante": int(saldo) if saldo else (int(qtd_contratada) if qtd_contratada else None),
            }

    return sorted(laboratorios), list(exames.values())


def main():
    engine = escolher_banco()
    Session = sessionmaker(bind=engine)
    sessao = Session()

    print("Lendo planilhas...")
    mnemonicos = carregar_mnemonicos()
    laboratorios, exames = carregar_apoio()
    print(f"  {len(laboratorios)} laboratorios encontrados")
    print(f"  {len(exames)} exames unicos encontrados")

    # ---------- Laboratorios ----------
    print("\nCadastrando laboratorios...")
    mapa_laboratorio_id = {}

    for nome_lab in laboratorios:
        resultado = sessao.execute(
            text("SELECT id FROM laboratorios WHERE nome = :nome"),
            {"nome": nome_lab}
        ).fetchone()

        if resultado:
            mapa_laboratorio_id[nome_lab] = resultado[0]
            print(f"  Ja existe: {nome_lab}")
        else:
            resultado = sessao.execute(
                text("INSERT INTO laboratorios (nome) VALUES (:nome) RETURNING id"),
                {"nome": nome_lab}
            ).fetchone()
            mapa_laboratorio_id[nome_lab] = resultado[0]
            print(f"  Criado: {nome_lab}")

    sessao.commit()

    # ---------- Exames ----------
    print("\nCadastrando exames...")
    criados = 0
    pulados = 0

    for exame in exames:
        nome_upper = exame["nome"].upper()

        ja_existe = sessao.execute(
            text("SELECT id FROM exames WHERE UPPER(nome) = :nome"),
            {"nome": nome_upper}
        ).fetchone()

        if ja_existe:
            pulados += 1
            continue

        sigla = mnemonicos.get(nome_upper)
        laboratorio_id = mapa_laboratorio_id.get(exame["laboratorio_nome"])

        sessao.execute(
            text("""
                INSERT INTO exames (nome, sigla, laboratorio_id, quantidade_contratada, quantidade_restante, ativo)
                VALUES (:nome, :sigla, :laboratorio_id, :qtd_contratada, :qtd_restante, true)
            """),
            {
                "nome": exame["nome"],
                "sigla": sigla,
                "laboratorio_id": laboratorio_id,
                "qtd_contratada": exame["quantidade_contratada"],
                "qtd_restante": exame["quantidade_restante"],
            }
        )
        criados += 1

    sessao.commit()
    sessao.close()

    print(f"\nConcluido!")
    print(f"  {criados} exames criados")
    print(f"  {pulados} exames ja existiam (pulados)")


if __name__ == "__main__":
    main()
